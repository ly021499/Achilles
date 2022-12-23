# @Time   : 2022/11/2 21:05
# @Author : LOUIE
# @Desc   : 解析excel数据

from utils.match import match_string
import time
from openpyxl import load_workbook
from utils.excel_tool import ExcelTool
import os
import setting


EXCEL_PATH = 'D:\\MainTrunk\\Excels\\Dev'
mapping_path = os.path.join(setting.RES_DIR, 'excel\\mapping.xlsx')
drop_bag_path = os.path.join(EXCEL_PATH, 'CfgDropBag.xlsx')
equipment_path = os.path.join(EXCEL_PATH, 'CfgEquipment.xlsx')
language_path = os.path.join(EXCEL_PATH, 'CfgLanguage.xlsx')
item_path = os.path.join(EXCEL_PATH, 'CfgItem.xlsx')
dungeon_path = os.path.join(EXCEL_PATH, 'CfgDungeon.xlsx')


def get_workbook_instance(filepath):
    """
    获得EXCEL workbook操作对象
    """
    workbook = load_workbook(filepath, data_only=True)
    return workbook


def get_worksheet_instance(filepath, sheet_name):
    """
    获得EXCEL表中的sheet操作对象
    """
    workbook = get_workbook_instance(filepath)
    worksheet = workbook[sheet_name]
    return worksheet


def get_target_sheet_data(filepath, sheet_name, idx, column_key, column_value, is_parser=True):
    """
    公共方法
    获取指定excel中的指定sheet表中的指定列的数据
    """
    ws = get_worksheet_instance(filepath, sheet_name)
    ws_rows_max = ws.max_row
    idx = idx
    items = {}
    while idx <= ws_rows_max:
        item_key = ws.cell(row=idx, column=column_key).value
        item_value = ws.cell(row=idx, column=column_value).value
        if is_parser:
            items.update({item_key: match_string(item_value)})
        else:
            items.update({item_key: item_value})
        idx += 1
    return items


def get_equipment_identifier():
    """
    获取 CfgEquipment 工作簿中的 CfgEquipment 表的 穿戴物ID、穿戴物名称
    """
    return get_target_sheet_data(equipment_path, 'CfgEquipment', 2, 2, 8, is_parser=False)


def traverse_dict(target_dict, drop_items):
    mapping_drop_items = {}
    for target_key, target_items in target_dict.items():
        for target_item in target_items:
            for current_key, current_items in drop_items.items():
                if str(target_item) == str(current_key):
                    mapping_drop_items.update({target_key: current_items})

    return mapping_drop_items


def get_regular_mapping_drop_items(sheet_name):
    """
    从表中拿到对应的名称，再到 CfgDropBag.xlsx - 掉落母表 ，获取名称映射到的 DropItem
    """

    target_regular_drop_items = get_target_sheet_data(mapping_path, sheet_name, 2, 1, 2)
    drop_items = get_target_sheet_data(drop_bag_path, '掉落母表', 2, 2, 7)
    # 第一步：拿到我需要的数据表
    regular_mapping_drop_items = traverse_dict(target_regular_drop_items, drop_items)

    return regular_mapping_drop_items


def get_first_mapping_drop_items(sheet_name):
    """
    从表中拿到对应的名称，再到 CfgDropBag.xlsx - 掉落母表 ，获取名称映射到的 DropItem， 要到两张表里取
    """
    # 这里拿到的首通的id后，到对应的表取名称标志符，drop_items错了

    target_first_drop_items = get_target_sheet_data(mapping_path, sheet_name, 2, 1, 5)
    first_drop_items = get_target_sheet_data(item_path, 'CfgItem', 4, 2, 4, is_parser=False)
    equipment_identifier = get_equipment_identifier()

    first_mapping_drop_items_1 = traverse_dict(target_first_drop_items, first_drop_items)
    first_mapping_drop_items_2 = traverse_dict(target_first_drop_items, equipment_identifier)

    first_mapping_drop_items = dict(**first_mapping_drop_items_1, **first_mapping_drop_items_2)

    return first_mapping_drop_items


def get_mapping_equipment_ident(mapping_drop_items):
    """获取对应的装备标志符"""
    equipment_identifier = get_equipment_identifier()
    equipment_ident_dict = {}
    for level_name, equ_ids in mapping_drop_items.items():
        identifier_list = []
        for equ_id in equ_ids:
            for equipment_id, identifier in equipment_identifier.items():
                if str(equipment_id) == str(equ_id):
                    identifier_list.append(identifier)

        equipment_ident_dict.update({level_name: identifier_list})
    return equipment_ident_dict


def get_mapping_ch_name(equipment_ident_dict):
    """获取对应的中文名称"""
    equipment_ch_name = get_target_sheet_data(language_path, 'CfgLanSystem', 4, 2, 5, is_parser=False)
    mapping_ch_name_dict = {}
    for level_name, identifiers in equipment_ident_dict.items():
        ch_name_list = []
        if isinstance(identifiers, str):
            identifiers = [identifiers]
        for identifier in identifiers:
            for ident, ch_name in equipment_ch_name.items():
                if str(ident) == str(identifier):
                    ch_name_list.append(ch_name)

        mapping_ch_name_dict.update({level_name: ch_name_list})
    return mapping_ch_name_dict


def write_to_excel(sheet_name):
    """
    将数据写入excel
    """
    target_regular_drop_items = get_regular_mapping_drop_items(sheet_name)
    regular_equipment_ident_dict = get_mapping_equipment_ident(target_regular_drop_items)
    regular_ch_name_dict = get_mapping_ch_name(regular_equipment_ident_dict)
    regular_items_data_list = [i for i in regular_ch_name_dict.values()]

    target_first_drop_items = get_first_mapping_drop_items(sheet_name)
    first_ch_name_dict = get_mapping_ch_name(target_first_drop_items)

    wb = get_workbook_instance(mapping_path)

    ws = wb[sheet_name]
    ws_rows_max = ws.max_row
    idx = 2

    while idx <= ws_rows_max:
        ws.cell(row=idx, column=4).value = str(regular_items_data_list[idx-2])
        key = ws.cell(row=idx, column=1).value
        ws.cell(row=idx, column=6).value = str(first_ch_name_dict[key])
        idx += 1

    wb.save(mapping_path)


def get_reward_config(sheet_name: str):
    """
    读取奖励的配置数据
    """
    ws = get_worksheet_instance(mapping_path, sheet_name)
    ws_rows_max = ws.max_row
    idx = 2

    reward_dict = {}
    while idx <= ws_rows_max:
        level_name = ws.cell(row=idx, column=3).value
        regular_reward_list = eval(ws.cell(row=idx, column=4).value)
        first_reward_list = ws.cell(row=idx, column=6).value
        if first_reward_list:
            regular_reward_list.extend(eval(first_reward_list))
        reward_dict.update(
            {level_name: regular_reward_list}
        )
        idx += 1

    return reward_dict


def write_assert_result(sheet_name, target_level_name, result, reward):
    et = ExcelTool(mapping_path)
    ws = et.get_sheet(sheet_name)
    ws_rows_max = ws.max_row
    idx = 2

    import datetime
    today = datetime.date.today()

    ws.cell(row=1, column=7).value = f"断言结果({today})"
    ws.cell(row=1, column=8).value = f"战斗奖励({today})"

    while idx <= ws_rows_max:
        level_name = ws.cell(row=idx, column=3).value
        if level_name == target_level_name:
            ws.cell(row=idx, column=7).value = str(result)
            ws.cell(row=idx, column=8).value = str(reward)

        idx += 1

    et.workbook.save(mapping_path)


def get_target_reward_data(reward_dict: dict, level_name: str):

    for level, reward_dict in reward_dict.items():
        if level == level_name:
            regular_reward = reward_dict
            return regular_reward
    return []


def mapping_ident_dict(mapping_data, target_data):

    mapping_dict = {}
    for level_name, items_id in mapping_data.items():
        items_list = []
        for item_id in items_id:
            if int(item_id) == 1:
                items_list.append('金币')
            if int(item_id) == 5:
                items_list.append('钻石')
            for ident_id, reward_name in target_data.items():
                if str(ident_id) == str(item_id):
                    items_list.append(reward_name)

        mapping_dict.update({level_name: items_list})
    return mapping_dict


def get_row_shadow_data():

    wb = ExcelTool(dungeon_path)
    ws = wb.get_sheet('CfgDungeonLevel')

    wb2 = ExcelTool(mapping_path)
    ws2 = wb2.get_sheet('虚影殿堂')

    start_idx = 85
    end_idx = 173
    item = {}
    while start_idx <= end_idx:
        item_key = ws.cell(row=start_idx, column=3).value
        item_value = ws.cell(row=start_idx, column=15).value
        start_idx += 1
        if '废弃' in str(item_key):
            continue
        item.update({item_key: item_value})

    idx = 2
    for key, value in item.items():
        ws2.cell(row=idx, column=1).value = key
        ws2.cell(row=idx, column=6).value = value
        idx += 1

    wb2.workbook.save(mapping_path)

    return item


def mapping_shadow_drop_item(item: dict):
    wb = ExcelTool(drop_bag_path)
    ws = wb.get_sheet('掉落母表')
    start_idx = 1650
    end_idx = 1739
    items2 = {}
    while start_idx <= end_idx:
        item_key = ws.cell(row=start_idx, column=2).value
        item_value = ws.cell(row=start_idx, column=7).value
        for key, value in item.items():
            if str(value) == str(item_key):
                items2.update({key: item_value})
        start_idx += 1

    wb2 = ExcelTool(mapping_path)
    ws2 = wb2.get_sheet('虚影殿堂')
    idx = 2
    for k, v in items2.items():
        ws2.cell(row=idx, column=5).value = v
        idx += 1

    wb2.workbook.save(mapping_path)
    return items2


def write_data(mapping_dict, sheet_name, column=4, idx=2):
    items_list = [i for i in mapping_dict.values()]
    wb = ExcelTool(mapping_path)
    ws = wb.get_sheet(sheet_name)
    ws_rows_max = ws.max_row
    while idx <= ws_rows_max:
        ws.cell(row=idx, column=column).value = str(items_list[idx-2])
        idx += 1
    wb.workbook.save(mapping_path)


def write_shadow_data():
    sheet_name = '虚影殿堂'
    mapping_data = get_target_sheet_data(mapping_path, sheet_name, 2, 1, 2, True)
    target_data = get_target_sheet_data(item_path, 'CfgItem', 4, 2, 3, False)
    mapping_dict = mapping_ident_dict(mapping_data, target_data)
    write_data(mapping_dict, sheet_name, 4)


def write_forbid_data():
    sheet_name = '贪婪禁地'
    mapping_data = get_target_sheet_data(mapping_path, sheet_name, 2, 1, 2, True)
    target_data = get_target_sheet_data(item_path, 'CfgItem', 4, 2, 3, False)
    mapping_dict = mapping_ident_dict(mapping_data, target_data)
    write_data(mapping_dict, sheet_name, 4)


def write_sky_data():
    sheet_name = '苍穹之城'
    mapping_data = get_target_sheet_data(mapping_path, sheet_name, 2, 1, 2, True)
    target_data = get_target_sheet_data(item_path, 'CfgItem', 4, 2, 4, False)
    mapping_dict = mapping_ident_dict(mapping_data, target_data)
    language_data = get_target_sheet_data(language_path, 'CfgLanItem', 4, 2, 4, False)

    dic_ = {}
    for k, v in mapping_dict.items():
        for j in v:
            for i, o in language_data.items():
                if str(i) == str(j):
                    v.remove(j)
                    v.append(o)
        dic_.update({k: v})

    dic_2 = {}
    for k, v in mapping_dict.items():
        for j in v:
            for i, o in language_data.items():
                if str(i) == str(j):
                    v.remove(j)
                    v.append(o)
        dic_2.update({k: v})

    write_data(dic_2, sheet_name, 4)


def init_reward_data():
    """
    初始化mapping表中的奖励数据
    """
    start_time = time.time()

    # write_to_excel('元素峡谷')
    write_to_excel('茫然遗迹')
    # write_shadow_data()
    # write_sky_data()
    # write_forbid_data()

    end_time = time.time()
    duration = round(end_time - start_time, 2)
    print('duration: ', duration)


