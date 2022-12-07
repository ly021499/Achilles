# @Time   : 2022/11/2 21:05
# @Author : LOUIE
# @Desc   : 解析excel数据

import re
import time
from openpyxl import load_workbook
import os
import setting


target_drop_bag_path = os.path.join(setting.RES_DIR, 'excel\\mapping.xlsx')
drop_bag_path = 'D:\\MainTrunk\\Excels\\Dev\\CfgDropBag.xlsx'
equipment_path = 'D:\\MainTrunk\\Excels\\Dev\\CfgEquipment.xlsx'
language_path = 'D:\\MainTrunk\\Excels\\Dev\\CfgLanguage.xlsx'
item_path = 'D:\\MainTrunk\\Excels\\Dev\\CfgItem.xlsx'


def parser_string(string: str):
    """
    正则匹配item的id
    """
    items = re.findall(r'\{(\d+?)\,', string)
    return items


def get_workbook_instance(excel_path):
    """
    获得EXCEL workbook操作对象
    """
    workbook = load_workbook(excel_path, data_only=True)
    return workbook


def get_worksheet_instance(excel_path, sheet_name):
    """
    获得EXCEL表中的sheet操作对象
    """
    workbook = get_workbook_instance(excel_path)
    worksheet = workbook[sheet_name]
    return worksheet


def get_target_sheet_data(excel_path, sheet_name, idx, column_key, column_value, is_parser=True):
    """
    公共方法
    获取指定excel中的指定sheet表中的指定列的数据
    """
    ws = get_worksheet_instance(excel_path, sheet_name)
    ws_rows_max = ws.max_row
    idx = idx
    items = {}
    while idx <= ws_rows_max:
        item_key = ws.cell(row=idx, column=column_key).value
        item_value = ws.cell(row=idx, column=column_value).value
        if is_parser:
            items.update({item_key: parser_string(item_value)})
        else:
            items.update({item_key: item_value})
        idx += 1
    return items


def get_first_drop_items():
    """
    获取 CfgItem 工作簿中的 CfgItem 中的 drop_item_id 和 drop_item数据
    """
    return get_target_sheet_data(item_path, 'CfgItem', 4, 2, 4, is_parser=False)


def get_drop_items():
    """
    获取 CfgDropBag 工作簿中的 掉落母表 中的 drop_item_id 和 drop_item数据
    """
    return get_target_sheet_data(drop_bag_path, '掉落母表', 2, 2, 7)


def get_target_regular_drop_item(sheet_name):
    """
    获取常规掉落物品items
    """
    return get_target_sheet_data(target_drop_bag_path, sheet_name, 2, 1, 5)


def get_target_first_drop_item(sheet_name):
    """
    获取首通掉落物品items
    """
    return get_target_sheet_data(target_drop_bag_path, sheet_name, 2, 1, 9)


def get_equipment_identifier():
    """
    获取 CfgEquipment 工作簿中的 CfgEquipment 表的 穿戴物ID、穿戴物名称
    """
    return get_target_sheet_data(equipment_path, 'CfgEquipment', 2, 2, 8, is_parser=False)


def get_item_ch_language():
    """
    获取 CfgLanguage 工作簿中的 CfgLanSystem 表的 序号、中文
    """
    return get_target_sheet_data(language_path, 'CfgLanSystem', 4, 2, 5, is_parser=False)


def traverse_dict(target_dict, drop_items):
    mapping_drop_items = {}
    for target_key, target_items in target_dict.items():
        for target_item in target_items:
            for current_key, current_items in drop_items.items():
                if str(target_item) == str(current_key):
                    mapping_drop_items.update({target_key: current_items})

    return mapping_drop_items


def get_regular_mapping_drop__items(sheet_name):
    """
    从表中拿到对应的名称，再到 CfgDropBag.xlsx - 掉落母表 ，获取名称映射到的 DropItem
    """

    target_regular_drop_items = get_target_regular_drop_item(sheet_name)
    drop_items = get_drop_items()
    # 第一步：拿到我需要的数据表
    regular_mapping_drop_items = traverse_dict(target_regular_drop_items, drop_items)

    return regular_mapping_drop_items


def get_first_mapping_drop_items(sheet_name):
    """
    从表中拿到对应的名称，再到 CfgDropBag.xlsx - 掉落母表 ，获取名称映射到的 DropItem， 要到两张表里取
    """
    # 这里拿到的首通的id后，到对应的表取名称标志符，drop_items错了

    target_first_drop_items = get_target_first_drop_item(sheet_name)
    first_drop_items = get_first_drop_items()
    equipment_identifier = get_equipment_identifier()

    first_mapping_drop_items_1 = traverse_dict(target_first_drop_items, first_drop_items)
    first_mapping_drop_items_2 = traverse_dict(target_first_drop_items, equipment_identifier)

    first_mapping_drop_items = dict(**first_mapping_drop_items_1, **first_mapping_drop_items_2)

    return first_mapping_drop_items


def get_mapping_equipment_ident(mapping_drop_items):
    equipment_identifier = get_equipment_identifier()
    equipment_ident_dict = {}
    for level_name, equ_ids in mapping_drop_items.items():
        identifier_list = []
        for equ_id in equ_ids:
            for equipment_id, identifier in equipment_identifier.items():
                if str(equipment_id) == str(equ_id):
                    identifier_list.append(identifier)

        equipment_ident_dict.update({level_name: identifier_list})
    return equipment_ident_dict


def get_mapping_ch_name(equipment_ident_dict):
    equipment_ch_name = get_item_ch_language()
    mapping_ch_name_dict = {}
    for level_name, identifiers in equipment_ident_dict.items():
        ch_name_list = []
        if isinstance(identifiers, str):
            identifiers = [identifiers]
        for identifier in identifiers:
            for ident, ch_name in equipment_ch_name.items():
                if str(ident) == str(identifier):
                    ch_name_list.append(ch_name)

        mapping_ch_name_dict.update({level_name: ch_name_list})
    return mapping_ch_name_dict


def write_to_excel(sheet_name):
    """
    将数据写入excel
    """
    # target_regular_drop_items = get_regular_mapping_drop__items(sheet_name)
    # regular_equipment_ident_dict = get_mapping_equipment_ident(target_regular_drop_items)
    # regular_ch_name_dict = get_mapping_ch_name(regular_equipment_ident_dict)
    # regular_items_data_list = [i for i in regular_ch_name_dict.values()]

    target_first_drop_items = get_first_mapping_drop_items(sheet_name)
    first_ch_name_dict = get_mapping_ch_name(target_first_drop_items)

    wb = get_workbook_instance(target_drop_bag_path)

    ws = wb[sheet_name]
    ws_rows_max = ws.max_row
    idx = 2

    while idx <= ws_rows_max:
        # ws.cell(row=idx, column=8).value = str(regular_items_data_list[idx-2])
        key = ws.cell(row=idx, column=1).value
        ws.cell(row=idx, column=10).value = str(first_ch_name_dict[key])
        idx += 1

    wb.save(target_drop_bag_path)


def get_reward_config(sheet_name: str):
    """
    读取奖励的配置数据
    """
    ws = get_worksheet_instance(target_drop_bag_path, sheet_name)
    ws_rows_max = ws.max_row
    idx = 2

    reward_dict = {}
    while idx <= ws_rows_max:
        level_name = ws.cell(row=idx, column=7).value
        regular_reward_list = ws.cell(row=idx, column=8).value
        first_reward_list = ws.cell(row=idx, column=10).value
        data = {
            level_name: {
                'regular': regular_reward_list,
                'first': first_reward_list
            }
        }
        reward_dict.update(data)

        idx += 1

    return reward_dict


def get_target_reward_data(reward_dict: dict, level_name: str):

    for level, reward_dict in reward_dict.items():
        if level == level_name:
            regular_reward = eval(reward_dict['regular'])
            regular_reward.extend(eval(reward_dict['first']))
            return regular_reward
    return []


# 分成三个步骤：1.初始化表数据，写入表数据后，存储 2.启动程序后，直接读取表中数据存储为生成器或者字典映射对象 3.每次要拿数据再去生成器中去取就可以了。


def init_reward_data():
    """
    初始化mapping表中的奖励数据
    """
    start_time = time.time()
    write_to_excel('元素峡谷')
    # write_to_excel('贪婪禁地')
    # write_to_excel('茫然遗迹')
    # write_to_excel('苍穹之城')
    # write_to_excel('虚影殿堂')
    end_time = time.time()
    duration = round(end_time - start_time, 2)
    print('duration: ', duration)


def struct_reward_data():
    """
    存储数据为对象
    """
    pass


def get_reward_data():
    """
    获取指定数据对象
    """
    pass


if __name__ == '__main__':
    # init_reward_data()
    print(get_target_reward_data(get_reward_config('茫然遗迹'), '比娜16'))




