# @Time   : 2022/11/2 21:05
# @Author : LOUIE
# @Desc   : excel操作工具类
import openpyxl
from utils.match import match_string


class ExcelTool:
    """
    封装一个处理excel文件的工具类
    """

    def __init__(self, filepath):
        """
        加载文件, 创建一个workbook对象
        :param filepath, excel文件路径
        """
        self.workbook = openpyxl.load_workbook(filepath, data_only=True)
        self.worksheet = None

    def get_sheet(self, var):
        """
        通过索引或者名字获取sheet
        :param var: 索引or名字
        :return: sheet
        """
        if isinstance(var, int):
            self.worksheet = self.workbook[self.workbook.sheetnames[var]]
            return self.worksheet

        if isinstance(var, str):
            self.worksheet = self.workbook[var]
            return self.worksheet

    def get_cell_value(self, row, col):
        """
        获取cell的值
        :param row: cell所在行
        :param col: cell所在列
        :return: cell的值
        """
        try:
            return self.worksheet.cell(row=row, column=col).value
        except Exception as e:
            return None

    def get_col_list(self, col, start_row=2, end_row=None):
        return self.get_parser_col_list( col, start_row=start_row, end_row=end_row, is_parser_value=False)

    def get_parser_col_list(self, col, start_row=2, end_row=None, is_parser_value=True):
        """
        获取指定列的所有数据
        :param col: 指定列, 如'A'
        :param start_row: 指定列, 如'A'
        :param end_row: 指定列, 如'A'
        :param is_parser_value: 指定列, 如'A'
        :return: 所有列数据组成的列表
        """

        if end_row is None:
            max_row = self.get_max_row()
        else:
            max_row = end_row

        if start_row > max_row:
            return None

        col_data_list = []
        while start_row <= max_row:
            var = self.get_cell_value(row=start_row, col=col)
            if is_parser_value is True:
                col_data_list.append(match_string(var))
            else:
                col_data_list.append(var)
            start_row += 1
        return col_data_list

    def get_col_dict(self, col_key, col_var, start_row=2, end_row=None, is_parser_value=False):
        col_key = self.get_col_list(col_key, start_row, end_row)
        col_var = self.get_col_list(col_var, start_row, end_row, is_parser_value=is_parser_value)
        col_dict = dict(zip(col_key, col_var))
        return col_dict

    def get_max_row(self):
        """
        获取最大行数
        :return: 最大行数
        """
        return self.worksheet.max_row

    def get_min_row(self):
        """
        获取最小行数
        :return: 最小行数
        """
        return self.worksheet.min_row

    def get_all_value(self):
        """
        获取所有数据
        :return:
        """
        return tuple(self.worksheet.values)

    def get_row_data(self, row):
        """
        获取指定行的所有数据，起始行为0
        :param row: 指定行
        :return: 所有行数据组成的列表
        """
        return list(self.get_all_value()[row])

    def write_data(self, row, col, value, path):
        """
        写入数据
        :param row: 指定cell所在行
        :param col: 指定cell所在列
        :param value: cell的值
        :param path: 保存文件路径
        """
        try:
            self.worksheet = self.workbook.active
            self.worksheet.cell(column=col, row=row, value=value)
            self.workbook.save(path)
        except BaseException as e:
            print(e)
            return None


if __name__ == '__main__':
    import os
    import setting
    excel_path = mapping_path = os.path.join(setting.RES_DIR, 'excel\\mapping.xlsx')
    wb = ExcelTool(excel_path)
    ws = wb.get_sheet('贪婪禁地')
    print(wb.get_col_dict(1, 5, is_parser_value=True))
